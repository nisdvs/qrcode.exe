import time, serial, sys, os, cv2
from tkinter import Tk, Frame, Label, font, PhotoImage, TOP
from scipy import *
from numpy import array
from PIL import Image, ImageTk, ImageOps
from pyzbar.pyzbar import decode
import json
import locale
from datetime import datetime
import pandas as pd
import time
from openpyxl import load_workbook


alunos_entrada = []

locale.setlocale(locale.LC_ALL, "pt_BR")
diaPT = datetime.now().strftime('%d/%m/%Y')
today = datetime.now().strftime('%a')

camera_id = 0
delay = 1
window_name = 'OpenCV pyzbar'

alunos_dict = {}
alunos_img = {}
alunos_data = pd.read_csv('assets/Lista.csv').to_dict()

for index in range(len(alunos_data.get('nome'))):
    alunos_dict[alunos_data['nome'][index]] = alunos_data['dias'][index].split(',') 
    alunos_img[alunos_data['nome'][index]] = alunos_data['imagem'][index]

print(alunos_dict.values())
print(alunos_img.values())

mGui = Tk()
mGui.title("Sistema de Gerenciamento de Almoço")
mGui.iconbitmap("assets/icon.ico")
largura = 1366
altura = 768
mGui.geometry('%dx%d+0+0' % (largura, altura))
mGui.configure(background="#892020")
mGui.state('zoomed')

titulo = Label(mGui, text="APONTE O QRCODE NA CÂMERA", width=40, height=2, bg="#EEE8E8", fg="#000000", font=("arial 25 bold"))
titulo.pack(side=TOP, pady=40)
titulo.config(borderwidth=2, relief="solid")


almoco = 'NENHUM ALUNO DETECTADO'
v2 = Label(mGui, text=almoco, width=70, height=1, background='#EEE8E8', foreground='black')
v2.config(font=('Arial', 20, 'bold'))
v2.pack(side=TOP, pady=10) 
v2.config(borderwidth=2, relief="solid")

blocoGeral = Frame(mGui, width=1070, height=380, bg='#892020')
blocoGeral.pack(side=TOP, pady=60)

camFrame = Frame(blocoGeral, width=370, height=370)
camFrame.place(x=0, y=0) #TROCAR LOCAL DO BLOCO
camFrame.config(borderwidth=2, relief="solid")
infoFrame = Frame(blocoGeral, width=370, height=370, bg='#EEE8E8')
infoFrame.place(x=700, y=0) #TROCAR LOCAL DO BLOCO
infoFrame.config(borderwidth=2, relief="solid")

cap = cv2.VideoCapture(camera_id)
ret, frame = cap.read()
v1 = Label(camFrame, width=362, height=362,text="QrCode Video")
v1.place(x=-0, y=-0)

def dddd():
    ret, frame = cap.read()
    global almoco
    if ret:
        for d in decode(frame):
            try:
                s = d.data.decode()
                json_data = json.loads(s)
                nome_aluno = json_data['nome']
                dias_almoco = json_data['comida']
                almoco = f'{nome_aluno} NÃO AUTORIZADO(A)!'

                #PROCURAR SE TEM O NOME NO DIC, SE O TODAY TEM NA LISTA DO QRCODE E SE O A LISTA DO QRCODE É IGUAL A DA PLANILHA
                if nome_aluno in alunos_dict.keys() and today in dias_almoco and dias_almoco == alunos_dict[nome_aluno]:
                    
                    
                    #PROCURAR SE O ALUNO ESTÁ NA PLANILHA
                    wb = load_workbook("assets/relacao.xlsx")
                    ws = wb.active
                    aluno_presente = False

                    for row in ws.iter_rows(values_only=True):
                        if nome_aluno in row and diaPT in row:
                            aluno_presente = True
                            break

                    #VERIFICAR SE O ALUNO JA PASSOU OU NAO
                    if not aluno_presente:
                        #ADD ALUNO NO EXCEL
                        # Carregar o arquivo do Excel
                        wb = load_workbook("assets/relacao.xlsx")

                        # Obter a primeira planilha (índice 0)
                        ws = wb.worksheets[0]

                        # Adicionar novo aluno ao Excel
                        new_row_data = [[nome_aluno, today, diaPT]]
                        for row_data in new_row_data:
                            ws.append(row_data)

                        # Salvar o arquivo Excel
                        wb.save("assets/relacao.xlsx")


                        #ALUNO LIBERADO E IMAGEM
                        almoco = f'{nome_aluno} AUTORIZADO(A)!'
                        v2.config(foreground='green') 
                        imagem_aluno =  alunos_img[nome_aluno]
                        imagem = Image.open('assets/fotos/' + imagem_aluno)
                        largura_desejada = 350
                        altura_desejada = 350
                        novo_tamanho = (largura_desejada, altura_desejada) 
                        imagem_redimensionada = imagem.resize(novo_tamanho)

                        borda_colorida = (0, 255, 0) 
                        largura_borda = 7
                        imagem_com_borda = ImageOps.expand(imagem_redimensionada, border=largura_borda, fill=borda_colorida)

                        # Converter a imagem com borda para PhotoImage
                        foto_aluno = ImageTk.PhotoImage(imagem_com_borda)

                        label_foto_aluno = Label(infoFrame, image=foto_aluno)
                        label_foto_aluno.image = foto_aluno 
                        image_height = imagem.height 
                        label_foto_aluno.place(x=0, y=0)  

                        #CARREGAR QRCODE
                        time.sleep(2);
                    
                    else:
                        almoco = f'{nome_aluno} JÁ FOI LIBERADO!'
                        v2.config(foreground='red')
                        imagem_aluno =  alunos_img[nome_aluno]
                        imagem = Image.open('assets/fotos/' + imagem_aluno)
                        largura_desejada = 350
                        altura_desejada = 350
                        novo_tamanho = (largura_desejada, altura_desejada) 
                        imagem_redimensionada = imagem.resize(novo_tamanho)
                        
                        borda_colorida = (255, 0, 0) 
                        largura_borda = 7
                        imagem_com_borda = ImageOps.expand(imagem_redimensionada, border=largura_borda, fill=borda_colorida)

                        # Converter a imagem com borda para PhotoImage
                        foto_aluno = ImageTk.PhotoImage(imagem_com_borda)

                        label_foto_aluno = Label(infoFrame, image=foto_aluno)
                        label_foto_aluno.image = foto_aluno 
                        image_height = imagem.height 
                        label_foto_aluno.place(x=0, y=0) 
                              
                else:
                    almoco = f'{nome_aluno} NÃO AUTORIZADO(A)!'
                    v2.config(foreground='red') 
                    imagem_aluno = alunos_img[nome_aluno]
                    imagem = Image.open('assets/fotos/' + imagem_aluno)
                    largura_desejada = 350
                    altura_desejada = 350
                    novo_tamanho = (largura_desejada, altura_desejada) 
                    imagem_redimensionada = imagem.resize(novo_tamanho)
                    
                    borda_colorida = (255, 0, 0) 
                    largura_borda = 7
                    imagem_com_borda = ImageOps.expand(imagem_redimensionada, border=largura_borda, fill=borda_colorida)

                    # Converter a imagem com borda para PhotoImage
                    foto_aluno = ImageTk.PhotoImage(imagem_com_borda)
                    
                    label_foto_aluno = Label(infoFrame, image=foto_aluno)
                    label_foto_aluno.image = foto_aluno
                    image_height = imagem.height 
                    label_foto_aluno.place(x=0, y=0)  
                    
            except Exception as e:
                print(e)
                pass
            
    img = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_RGB2BGR))
    nimg = ImageTk.PhotoImage(image=img)
    
   

    image = Image.open("assets/sesiaa.png")
    image.thumbnail((150, 150))
    photo = ImageTk.PhotoImage(image)


    image_label = Label(blocoGeral, image=photo)
    image_label.image = photo 


    image_height = image.height 

    image_label.place(x=460, y=200 - image_height)   


    
    v1.n_img = nimg
    v1.configure(image=nimg)
    v2.config(text=almoco)
    
    mGui.after(10, dddd)    
dddd()
mGui.mainloop()